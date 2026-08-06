# -*- coding: utf-8 -*-
"""
Restaurant Survival Toolkit 2026 — Saudi F&B Financial System
The 8 book chapters translated into 5 interactive Excel hubs.
PREMIUM LIGHT EXECUTIVE THEME:
  - White / cream (#F8F9FA) backgrounds, no dark panels.
  - Luxury Navy (#1B365D) + Bronze (#996515) headers.
  - 12pt pitch-black text everywhere (no gray text).
  - Gold-tint (#FFF2CC) instruction boxes with bold black text.
  - Thick clean borders separate input vs calculated blocks.
2026 verified rates: VAT 15%, Zakat 2.5%, GOSI 11.75/9.75 Saudi
(12.75/10.75 new system from Jul 2026), expat 2% employer,
Iqama 650/yr + work permit levy ~800/mo, CHI ~100/mo.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.protection import Protection
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule, CellIsRule

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "public", "Restaurant-Survival-Toolkit-2026.xlsx")

# ---------- Premium Light Executive palette ----------
NAVY        = "1B365D"   # luxury navy
NAVY_DARK   = "16305A"
BRONZE      = "996515"   # luxury gold/bronze
GOLD_TINT   = "FFF2CC"   # soft yellow/gold box
INPUT_FILL  = "FFF9E8"   # pale gold-tint editable cells
RESULT_FILL = "F2F2F2"   # light gray computed cells
ALT_FILL    = "F8F9FA"   # ultra-light gray zebra
WHITE       = "FFFFFF"
BLACK       = "000000"

# ---------- Fonts (Calibri, 12pt base, black) ----------
F_TITLE = Font(name="Calibri", size=18, bold=True, color=NAVY)
F_SUB   = Font(name="Calibri", size=12, color=BLACK)
F_SEC   = Font(name="Calibri", size=14, bold=True, color=NAVY)
F_HDR   = Font(name="Calibri", size=12, bold=True, color=WHITE)
F_LBL   = Font(name="Calibri", size=12, color=BLACK)
F_LBL_B = Font(name="Calibri", size=12, bold=True, color=BLACK)
F_IN    = Font(name="Calibri", size=12, bold=True, color=BLACK)
F_RES   = Font(name="Calibri", size=12, bold=True, color=BLACK)
F_NOTE  = Font(name="Calibri", size=12, bold=True, color=BLACK)
F_SMALL = Font(name="Calibri", size=11, color=BLACK)

# ---------- Borders ----------
S_THIN   = Side(style="thin", color="B7B7B7")
S_MED    = Side(style="medium", color=NAVY)
S_MED_G  = Side(style="medium", color=BRONZE)
B_THIN_ALL = Border(left=S_THIN, right=S_THIN, top=S_THIN, bottom=S_THIN)
B_INPUT    = Border(left=S_MED_G, right=S_MED_G, top=S_MED_G, bottom=S_MED_G)
B_RES      = Border(left=S_THIN, right=S_THIN, top=Side(style="medium", color=NAVY), bottom=S_THIN)
B_ROW      = Border(bottom=S_THIN)
B_BOX      = Border(left=S_MED_G, right=S_MED_G, top=S_MED_G, bottom=S_MED_G)
B_HEAD     = Border(bottom=Side(style="medium", color=NAVY_DARK))
B_SECTION  = Border(top=Side(style="medium", color=BRONZE), bottom=Side(style="medium", color=BRONZE))

# ---------- Fills ----------
FILL_INPUT  = PatternFill("solid", fgColor=INPUT_FILL)
FILL_RES    = PatternFill("solid", fgColor=RESULT_FILL)
FILL_HDR    = PatternFill("solid", fgColor=NAVY)
FILL_SECT   = PatternFill("solid", fgColor=GOLD_TINT)
FILL_HOWTO  = PatternFill("solid", fgColor=GOLD_TINT)
FILL_ALT    = PatternFill("solid", fgColor=ALT_FILL)
FILL_WHITE  = PatternFill("solid", fgColor=WHITE)

AL_R   = Alignment(horizontal="right", vertical="center")
AL_RW  = Alignment(horizontal="right", vertical="center", wrap_text=True)
AL_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_L   = Alignment(horizontal="left", vertical="center")

MONTHS = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
          "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]


def title_block(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(1, 1, title); c.font = F_TITLE; c.alignment = AL_C
    ws.row_dimensions[1].height = 36
    c2 = ws.cell(2, 1, sub); c2.font = F_SUB; c2.alignment = AL_C
    c2.border = Border(bottom=Side(style="medium", color=BRONZE))
    ws.row_dimensions[2].height = 24


def section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font = F_SEC; c.fill = FILL_SECT; c.alignment = AL_R
    c.border = B_SECTION
    ws.row_dimensions[row].height = 28


def header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = F_HDR; c.fill = FILL_HDR; c.alignment = AL_C
        c.border = B_HEAD
    ws.row_dimensions[row].height = 30


def label_cell(ws, row, col, text, bold=False, alt=False):
    c = ws.cell(row, col, text)
    c.font = F_LBL_B if bold else F_LBL
    c.fill = FILL_ALT if alt else FILL_WHITE
    c.alignment = AL_RW
    c.border = B_ROW
    return c


def input_cell(ws, row, col, value=None, fmt=None):
    c = ws.cell(row, col, value)
    c.fill = FILL_INPUT; c.font = F_IN; c.border = B_INPUT
    c.protection = Protection(locked=False)
    c.alignment = AL_C
    if fmt:
        c.number_format = fmt
    return c


def res_cell(ws, row, col, value=None, fmt=None):
    """Computed/result cell — light gray fill, bold black, clearly separated."""
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    c.fill = FILL_RES; c.font = F_RES; c.alignment = AL_C
    c.border = B_RES
    if fmt:
        c.number_format = fmt
    return c


def lock_all(ws, max_row, max_col, unlock=()):
    for r in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(r, col)
            cell.protection = (Protection(locked=False)
                               if (r, col) in unlock
                               else Protection(locked=True, hidden=False))
    ws.protection.sheet = True
    ws.protection.enable()


def widths(ws, wl):
    for i, w in enumerate(wl, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def autofit(ws, min_w=11, max_w=46, pad=3.5):
    """Auto-fit columns from static (non-formula, non-merged) text with padding.
    Only ever grows existing widths — never shrinks them."""
    merged_anchors = {rng.min_row * 10000 + rng.min_col for rng in ws.merged_cells.ranges}
    for col in range(1, ws.max_column + 1):
        best = 0
        for row in range(1, ws.max_row + 1):
            key = row * 10000 + col
            if key in merged_anchors:
                continue
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            v = str(cell.value)
            if v.startswith("="):
                continue
            w = 0
            for ch in v:
                if ch.isdigit() or ch in ".%,-$()' ":
                    w += 1.0
                elif ord(ch) > 0x0600:
                    w += 1.25
                else:
                    w += 0.95
            if v:
                best = max(best, w)
        if best:
            cur = ws.column_dimensions[get_column_letter(col)].width or 0
            ws.column_dimensions[get_column_letter(col)].width = max(cur, min(max_w, max(min_w, best + pad)))


def base_setup(ws, ncols):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


wb = Workbook()
wb.remove(wb.active)
wb.properties.title = "Restaurant Survival Toolkit 2026"
wb.properties.creator = "Restaurant Survival Toolkit"

# ============================================================
# SHEET 0 — README
# ============================================================
ws0 = wb.create_sheet("دليل الاستخدام")
base_setup(ws0, 6)
title_block(ws0, "حقيبة بقاء المطاعم 2026 — Restaurant Survival Toolkit",
            "منهجية الكتاب الكامل (8 فصول) مترجمة إلى أدوات تفاعلية — مصممة للبيئة التنظيمية السعودية 2026", 6)

r = 4
section(ws0, r, "محتويات الحقيبة", 6); r += 1
rows0 = [
    ("لوحة الاستراتيجية", "احتمالية بقاء مطعمك من مدخلاتك المالية + رسم الإيرادات مقابل نقطة التعادل"),
    ("الامتثال والضرائب", "ضريبة القيمة المضافة 15%، الزكاة، والتكلفة الحقيقية للموظف (GOSI + Muqeem + تأمين CHI)"),
    ("التكلفة الأولية والمنيو", "تسعير الطبق بدقة + مصفوفة هندسة المنيو (نجوم / حصان الجر / لغز / كلاب)"),
    ("كشف التدفق النقدي", "كشف جنائي لـ 12 شهراً يكشف أين تختفي الأرباح + فحص التسريبات الخفية"),
    ("قوائم التشغيل", "قوائم فتح وإغلاق يومية تفاعلية جاهزة للتصدير PDF"),
]
for i, (t, d) in enumerate(rows0):
    c = label_cell(ws0, r, 1, t, bold=True, alt=(i % 2 == 1))
    c.border = B_THIN_ALL
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c2 = label_cell(ws0, r, 2, d, alt=(i % 2 == 1))
    c2.border = B_THIN_ALL
    ws0.row_dimensions[r].height = 26
    r += 1
r += 1

# ---- How-to box: BLACK BOLD on soft gold #FFF2CC ----
section(ws0, r, "طريقة الاستخدام", 6); r += 1
howto = [
    "1. أدخل أرقامك في الخلايا ذات الإطار البني/الذهبي فقط — بقية الخلايا مقفلة لحماية الصيغ.",
    "2. ابدأ من لوحة الاستراتيجية: أدخل مدخلات الشهر واقرأ احتمالية البقاء فوراً.",
    "3. معدلات 2026 مدمجة مسبقاً وقابلة للتعديل: ضريبة القيمة المضافة 15%، الزكاة 2.5%، GOSI.",
    "4. للتصدير PDF: من قائمة File → Print → Save as PDF (قوائم التشغيل مضبوطة للطباعة).",
]
for i, t in enumerate(howto):
    ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws0.cell(r, 1, t)
    c.font = F_NOTE; c.fill = FILL_HOWTO; c.alignment = AL_RW
    c.border = B_BOX
    ws0.row_dimensions[r].height = 30
    r += 1
r += 1

section(ws0, r, "مصادر المعدلات 2026", 6); r += 1
for s in [
    "GOSI: سعودي = 11.75% صاحب عمل + 9.75% موظف (النظام الجديد اعتباراً من يوليو 2026: 12.75% + 10.75%) — الوافد 2% فقط على صاحب العمل.",
    "رسوم الإقامة: تجديد الإقامة 650 ر.س سنوياً + رسوم رخصة عمل ~800 ر.س شهرياً.",
    "التأمين الصحي CHI: إلزامي ويدفعه صاحب العمل — الفئة الأساسية من ~100 ر.س شهرياً.",
    "ضريبة القيمة المضافة 15% والزكاة 2.5% وفق هيئة الزكاة والضريبة والجمارك ZATCA.",
]:
    ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws0.cell(r, 1, s); c.font = F_SMALL; c.alignment = AL_RW
    ws0.row_dimensions[r].height = 20
    r += 1
widths(ws0, [30, 26, 26, 26, 26, 26])
autofit(ws0)
lock_all(ws0, r, 6)
print("Sheet 0 done")

# ============================================================
# SHEET 1 — DASHBOARD (Ch1)
# ============================================================
ws1 = wb.create_sheet("لوحة الاستراتيجية")
base_setup(ws1, 8)
title_block(ws1, "لوحة الاستراتيجية — Survival Dashboard",
            "احتمالية بقاء مطعمك وفق مدخلاتك المالية — الفصل الأول", 8)

r = 4
section(ws1, r, "مدخلات الشهر — أدخل أرقامك في الخلايا ذات الإطار الذهبي", 8); r += 1
kpis = [
    ("الإيرادات الشهرية شاملة الضريبة (ر.س)", 320000),
    ("تكلفة المواد الأولية COGS (ر.س)", 105000),
    ("تكلفة العمالة (ر.س)", 88000),
    ("الإيجار والمصاريف الثابتة (ر.س)", 52000),
    ("مصاريف تشغيل متغيرة (ر.س)", 21000),
    ("مبيعات ذروة رمضان/الأعياد (ر.س)", 520000),
]
kpi_rows = {}
for i, (name, val) in enumerate(kpis):
    label_cell(ws1, r, 1, name, bold=True, alt=(i % 2 == 1))
    input_cell(ws1, r, 2, val, '0')
    kpi_rows[name] = r
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws1.cell(r, 3).fill = FILL_ALT
    ws1.row_dimensions[r].height = 26
    r += 1
rev_row, cogs_row, lab_row, fix_row, var_row, peak_row = (kpi_rows[n] for n, _ in kpis)
r += 1

section(ws1, r, "مؤشرات محسوبة", 8); r += 1
net_rev_r = r;     r += 1
prime_r   = r;     r += 1
prime_pct = r;     r += 1
op_prof_r = r;     r += 1
net_m_r   = r;     r += 1
bep_r     = r;     r += 1
safe_r    = r;     r += 1
vat_r     = r + 1  # row 21 reserved: VAT rate input

label_cell(ws1, net_rev_r, 1, "صافي الإيرادات بعد الضريبة (ر.س)", bold=True)
res_cell(ws1, net_rev_r, 2, f"=B{rev_row}/(1+$B${vat_r})", '0')
label_cell(ws1, prime_r, 1, "التكلفة الأولية Prime Cost (ر.س)", bold=True)
res_cell(ws1, prime_r, 2, f"=B{cogs_row}+B{lab_row}", '0')
label_cell(ws1, prime_pct, 1, "نسبة التكلفة الأولية", bold=True)
res_cell(ws1, prime_pct, 2, f"=B{prime_r}/B{net_rev_r}", '0.0%')
label_cell(ws1, op_prof_r, 1, "صافي الربح التشغيلي (ر.س)", bold=True)
res_cell(ws1, op_prof_r, 2, f"=B{net_rev_r}-B{prime_r}-B{fix_row}-B{var_row}", '0')
label_cell(ws1, net_m_r, 1, "هامش الربح الصافي", bold=True)
res_cell(ws1, net_m_r, 2, f"=B{op_prof_r}/B{net_rev_r}", '0.0%')
label_cell(ws1, bep_r, 1, "نقطة التعادل الشهرية (ر.س)", bold=True)
res_cell(ws1, bep_r, 2, f"=IF(B{net_rev_r}>0,(B{fix_row}+B{var_row})/(1-B{prime_pct}),0)", '0')
label_cell(ws1, safe_r, 1, "هامش الأمان (ر.س)", bold=True)
res_cell(ws1, safe_r, 2, f"=B{net_rev_r}-B{bep_r}", '0')

label_cell(ws1, vat_r, 1, "معدل ضريبة القيمة المضافة", bold=True)
input_cell(ws1, vat_r, 2, 0.15, '0%')

r = vat_r + 1
section(ws1, r, "احتمالية البقاء", 8); r += 1
score_r = r
label_cell(ws1, score_r, 1, "درجة البقاء (من 100)", bold=True)
score_f = (f"=MIN(100,MAX(0,"
           f"40*MIN(1,MAX(0,B{safe_r})/MAX(B{bep_r},1))"
           f"+30*MAX(0,1-B{prime_pct})"
           f"+30*MAX(0,B{net_m_r})))")
res_cell(ws1, score_r, 2, score_f, '0')
label_cell(ws1, score_r, 3, "التصنيف", bold=True)
res_cell(ws1, score_r, 4, f'=IF(B{score_r}>=75,"قوي جداً",IF(B{score_r}>=55,"مستقر",IF(B{score_r}>=35,"هش","خطر")))')
ws1.row_dimensions[score_r].height = 28

r += 1
label_cell(ws1, r, 1, "شريط الحالة")
bar = ws1.cell(r, 2, f'=REPT("█",ROUND(B{score_r}/10,0))')
bar.font = Font(name="Calibri", size=14, bold=True, color=BRONZE)
bar.fill = FILL_RES; bar.alignment = AL_L
bar.protection = Protection(locked=True, hidden=False)
ws1.row_dimensions[r].height = 24
r += 1
label_cell(ws1, r, 1, "الحالة")
res_cell(ws1, r, 2, f'=IF(B{safe_r}>0,"فوق نقطة التعادل","تحت نقطة التعادل")')
ws1.row_dimensions[r].height = 24
r += 1
label_cell(ws1, r, 1, "مقارنة بذروة رمضان/الأعياد")
res_cell(ws1, r, 2, f"=IF(B{peak_row}>0,B{net_rev_r}/B{peak_row},1)", '0%')
ws1.row_dimensions[r].height = 24
r += 2

section(ws1, r, "الإيرادات الشهرية مقابل نقطة التعادل (بيانات الرسم)", 8); r += 1
chart_hdr = r
header_row(ws1, chart_hdr, ["الشهر", "الإيرادات الشهرية (ر.س)", "نقطة التعادل (ر.س)"])
chart_start = r + 1
for i, m in enumerate(MONTHS):
    rr = chart_start + i
    c1 = ws1.cell(rr, 1, m)
    c1.font = F_LBL; c1.fill = FILL_WHITE if i % 2 == 0 else FILL_ALT
    c1.alignment = AL_R; c1.border = B_THIN_ALL
    c1.protection = Protection(locked=True, hidden=False)
    res_cell(ws1, rr, 2, f"=IF(MOD(ROW(),3)=0,B{peak_row},B{rev_row})", '0')
    res_cell(ws1, rr, 3, f"=B{bep_r}", '0')
    ws1.row_dimensions[rr].height = 24
chart_last = chart_start + 11

ch = LineChart()
ch.title = "الإيرادات الشهرية مقابل نقطة التعادل"
ch.style = 13; ch.height = 9; ch.width = 26
data = Reference(ws1, min_col=2, min_row=chart_hdr, max_col=3, max_row=chart_last)
cats = Reference(ws1, min_col=1, min_row=chart_start, max_row=chart_last)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.y_axis.title = "ر.س"
ch.x_axis.title = "الشهر"
ch.legend.position = "b"
try:
    ch.series[0].graphicalProperties.line.solidFill = NAVY
    ch.series[0].graphicalProperties.line.width = 24000
    ch.series[1].graphicalProperties.line.solidFill = BRONZE
    ch.series[1].graphicalProperties.line.width = 24000
except Exception:
    pass
ws1.add_chart(ch, "E7")

widths(ws1, [40, 22, 22, 20, 18, 16, 16, 16])
autofit(ws1)
lock_all(ws1, chart_last + 2, 8,
         unlock={(rev_row, 2), (cogs_row, 2), (lab_row, 2), (fix_row, 2), (var_row, 2),
                 (peak_row, 2), (vat_r, 2)})
print("Sheet 1 done")

# ============================================================
# SHEET 2 — COMPLIANCE & LABOR (Ch2/3/6)
# ============================================================
ws2 = wb.create_sheet("الامتثال والضرائب")
base_setup(ws2, 10)
title_block(ws2, "الامتثال والضرائب — التكلفة الحقيقية للموظف",
            "GOSI + Muqeem + تأمين CHI + ضريبة القيمة المضافة + الزكاة — الفصول 2 و3 و6", 10)

r = 4
section(ws2, r, "معدلات 2026 (قابلة للتعديل)", 10); r += 1
rates = [
    ("GOSI سعودي (نظام قديم) — اشتراك صاحب العمل", 0.1175),
    ("GOSI سعودي (نظام قديم) — اشتراك الموظف", 0.0975),
    ("GOSI سعودي (النظام الجديد) — اشتراك صاحب العمل", 0.1275),
    ("GOSI سعودي (النظام الجديد) — اشتراك الموظف", 0.1075),
    ("GOSI الوافد — اشتراك صاحب العمل (المخاطر المهنية)", 0.02),
    ("رسوم رخصة العمل الشهرية (ر.س)", 800),
    ("رسوم تجديد الإقامة السنوية (ر.س)", 650),
    ("التأمين الصحي الأساسي CHI شهرياً (ر.س)", 100),
]
rate_rows = {}
for i, (name, val) in enumerate(rates):
    label_cell(ws2, r, 1, name, bold=True, alt=(i % 2 == 1))
    input_cell(ws2, r, 2, val, '0.00%' if isinstance(val, float) else '0')
    rate_rows[name] = r
    ws2.row_dimensions[r].height = 24
    r += 1
label_cell(ws2, r, 1, "تفعيل نظام GOSI الجديد (نعم/لا)", bold=True)
toggle_r = r
input_cell(ws2, toggle_r, 2, "لا")
dv_toggle = DataValidation(type="list", formula1='"نعم,لا"', allow_blank=True)
ws2.add_data_validation(dv_toggle); dv_toggle.add(f"B{toggle_r}")
ws2.row_dimensions[r].height = 24
r += 1

r += 1
section(ws2, r, "التكلفة الحقيقية لكل موظف", 10); r += 1
emp_hdr = r
header_row(ws2, emp_hdr,
           ["الاسم", "الجنسية", "الراتب الأساسي", "البدلات", "إجمالي الأجر",
            "اشتراك الموظف GOSI", "اشتراك صاحب العمل GOSI", "التأمين الصحي CHI",
            "رسوم شهرية (إقامة + عمل)", "التكلفة الشهرية الكاملة"])
emp_start = r + 1
employees = [
    ("مدير المطعم", "سعودي", 12000, 1500),
    ("رئيس الطهاة", "وافد", 8500, 1000),
    ("طباخ", "وافد", 5000, 500),
    ("نادل", "سعودي", 4500, 300),
    ("مشرف قاعة", "وافد", 4000, 400),
]
for i, (name, nat, base, allow) in enumerate(employees):
    rr = emp_start + i
    label_cell(ws2, rr, 1, name, alt=(i % 2 == 1))
    input_cell(ws2, rr, 2, nat)
    input_cell(ws2, rr, 3, base, '0')
    input_cell(ws2, rr, 4, allow, '0')
    res_cell(ws2, rr, 5, f"=C{rr}+D{rr}", '0')
    res_cell(ws2, rr, 6, f'=IF($B${rr}="سعودي",E{rr}*IF($B${toggle_r}="نعم",$B$8,$B$6),0)', '0')
    res_cell(ws2, rr, 7, f'=IF($B${rr}="سعودي",E{rr}*IF($B${toggle_r}="نعم",$B$7,$B$5),E{rr}*$B$9)', '0')
    input_cell(ws2, rr, 8, 100, '0')
    res_cell(ws2, rr, 9, f'=IF($B${rr}="وافد",$B$10+$B$11/12,0)', '0')
    res_cell(ws2, rr, 10, f"=E{rr}+G{rr}+H{rr}+I{rr}", '0')
    ws2.row_dimensions[rr].height = 24
emp_last = emp_start + len(employees) - 1
tot_r = emp_last + 1
label_cell(ws2, tot_r, 1, "الإجمالي الشهري", bold=True)
for col in range(3, 11):
    cl = get_column_letter(col)
    res_cell(ws2, tot_r, col, f"=SUM({cl}{emp_start}:{cl}{emp_last})", '0')
ws2.row_dimensions[tot_r].height = 26
dv_nat = DataValidation(type="list", formula1='"سعودي,وافد"', allow_blank=True)
ws2.add_data_validation(dv_nat); dv_nat.add(f"B{emp_start}:B{emp_last}")
r = tot_r + 2

section(ws2, r, "الالتزامات الضريبية (VAT)", 10); r += 1
rev2 = r
sales_vat = r + 1
inp = r + 2
net_vat = r + 3
vat_rate2 = r + 4
label_cell(ws2, rev2, 1, "الإيرادات الشهرية شاملة الضريبة (ر.س)", bold=True)
input_cell(ws2, rev2, 2, 320000, '0')
label_cell(ws2, sales_vat, 1, "ضريبة المبيعات المحصلة (15%)", bold=True)
res_cell(ws2, sales_vat, 2, f"=B{rev2}*$B${vat_rate2}", '0')
label_cell(ws2, inp, 1, "ضريبة المشتريات المدفوعة (Input VAT)", bold=True)
input_cell(ws2, inp, 2, 18000, '0')
label_cell(ws2, net_vat, 1, "صافي ضريبة القيمة المضافة المستحقة", bold=True)
res_cell(ws2, net_vat, 2, f"=B{sales_vat}-B{inp}", '0')
label_cell(ws2, vat_rate2, 1, "معدل ضريبة القيمة المضافة", bold=True)
input_cell(ws2, vat_rate2, 2, 0.15, '0%')
for rr in (rev2, sales_vat, inp, net_vat, vat_rate2):
    ws2.row_dimensions[rr].height = 24
r = vat_rate2 + 2

section(ws2, r, "الزكاة (Zakat)", 10); r += 1
zbase = r
zrate = r + 1
zdue = r + 2
label_cell(ws2, zbase, 1, "صافي الربح السنوي (ر.س)", bold=True)
input_cell(ws2, zbase, 2, 480000, '0')
label_cell(ws2, zrate, 1, "معدل الزكاة", bold=True)
input_cell(ws2, zrate, 2, 0.025, '0.0%')
label_cell(ws2, zdue, 1, "الزكاة السنوية المستحقة", bold=True)
res_cell(ws2, zdue, 2, f"=B{zbase}*B{zrate}", '0')
for rr in (zbase, zrate, zdue):
    ws2.row_dimensions[rr].height = 24
r = zdue + 1

widths(ws2, [42, 18, 18, 18, 18, 18, 18, 18, 22, 22])
autofit(ws2)
unlock2 = {(toggle_r, 2), (rev2, 2), (inp, 2), (vat_rate2, 2), (zbase, 2), (zrate, 2)}
for rr in range(emp_start, emp_last + 1):
    unlock2 |= {(rr, c) for c in (2, 3, 4, 8)}
for rr, _ in rates:
    unlock2.add((rr, 2))
lock_all(ws2, r, 10, unlock=unlock2)
print("Sheet 2 done")

# ============================================================
# SHEET 3 — PLATE COSTING + MENU ENGINEERING (Ch4/7)
# ============================================================
ws3 = wb.create_sheet("التكلفة والمنيو")
base_setup(ws3, 8)
title_block(ws3, "التكلفة الأولية وهندسة المنيو — Plate Costing & Menu Engineering",
            "سعر الطبق، تكلفته، هامشه، وتصنيفه (نجم / حصان الجر / لغز / كلب) — الفصلان 4 و7", 8)

r = 4
section(ws3, r, "مصفوفة هندسة المنيو", 8); r += 1
hdr = r
header_row(ws3, hdr, ["الطبق", "سعر البيع (ر.س)", "تكلفة الطبق (ر.س)",
                      "الوحدات المباعة شهرياً", "هامش الوحدة", "إجمالي الهامش",
                      "نسبة تكلفة الطعام", "تصنيف المنيو"])
dishes = [
    ("كبسة لحم", 48, 16),
    ("مندي دجاج", 38, 14),
    ("مطازيز", 42, 18),
    ("سمك مشوي", 65, 30),
    ("مشويات مشكلة", 75, 28),
    ("مقبلات", 25, 12),
    ("حلويات", 18, 9),
]
units_demo = [320, 520, 200, 430, 470, 380, 280]
d_start = r + 1
d_last = d_start + len(dishes) - 1
s1, s2, s3r = d_last + 1, d_last + 2, d_last + 3
for i, (name, price, cost) in enumerate(dishes):
    rr = d_start + i
    label_cell(ws3, rr, 1, name, alt=(i % 2 == 1))
    input_cell(ws3, rr, 2, price, '0')
    input_cell(ws3, rr, 3, cost, '0')
    input_cell(ws3, rr, 4, units_demo[i], '0')
    res_cell(ws3, rr, 5, f"=B{rr}-C{rr}", '0')
    res_cell(ws3, rr, 6, f"=D{rr}*E{rr}", '0')
    res_cell(ws3, rr, 7, f"=IF(B{rr}>0,C{rr}/B{rr},0)", '0.0%')
    res_cell(ws3, rr, 8, f"=IF(AND(E{rr}>$B${s2},D{rr}>$B${s1}),\"نجم ★\",IF(AND(E{rr}<$B${s2},D{rr}>$B${s1}),\"حصان الجر\",IF(AND(E{rr}>$B${s2},D{rr}<$B${s1}),\"لغز\",\"كلب\")))")
    ws3.row_dimensions[rr].height = 26
label_cell(ws3, s1, 1, "متوسط الوحدات المباعة", bold=True)
res_cell(ws3, s1, 2, f"=AVERAGE(D{d_start}:D{d_last})", '0')
label_cell(ws3, s2, 1, "متوسط هامش الوحدة", bold=True)
res_cell(ws3, s2, 2, f"=AVERAGE(E{d_start}:E{d_last})", '0')
label_cell(ws3, s3r, 1, "متوسط نسبة تكلفة الطعام", bold=True)
res_cell(ws3, s3r, 2, f"=AVERAGE(G{d_start}:G{d_last})", '0.0%')
for rr in (s1, s2, s3r):
    ws3.row_dimensions[rr].height = 26
r = s3r + 2

section(ws3, r, "التوصيات", 8); r += 1
for note in [
    "النجم ★: هامش مرتفع وشعبية عالية — ركّز التسويق عليه واجعل منه المفضّل.",
    "حصان الجر: شعبي جداً لكن هامشه منخفض — ارفع سعره بحذر أو قلّل تكلفة مكوناته.",
    "اللغز: هامش ممتاز لكن مبيعاته ضعيفة — حسّن عرضه أو ادمجه في عروض.",
    "كلب: هامش ومبيعات منخفضان — أعد تسعيره أو احذفه من المنيو.",
]:
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws3.cell(r, 1, note); c.font = F_LBL; c.alignment = AL_RW
    ws3.row_dimensions[r].height = 22
    r += 1

widths(ws3, [24, 18, 18, 22, 18, 18, 22, 20])
autofit(ws3)
unlock3 = {(rr, c) for rr in range(d_start, d_last + 1) for c in (2, 3, 4)}
lock_all(ws3, r, 8, unlock=unlock3)
rng = f"H{d_start}:H{d_last}"
ws3.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("نجم",H{d_start}))'],
                                                fill=PatternFill("solid", fgColor="C6EFCE"),
                                                font=Font(name="Calibri", color="006100", bold=True)))
ws3.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("لغز",H{d_start}))'],
                                                fill=PatternFill("solid", fgColor="FFEB9C"),
                                                font=Font(name="Calibri", color="9C6500", bold=True)))
ws3.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("حصان",H{d_start}))'],
                                                fill=PatternFill("solid", fgColor="DDEBF7"),
                                                font=Font(name="Calibri", color="1F4E79", bold=True)))
ws3.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("كلب",H{d_start}))'],
                                                fill=PatternFill("solid", fgColor="FFC7CE"),
                                                font=Font(name="Calibri", color="9C0006", bold=True)))
print("Sheet 3 done")

# ============================================================
# SHEET 4 — FORENSIC CASH FLOW (Ch5/8)
# ============================================================
ws4 = wb.create_sheet("كشف التدفق النقدي")
base_setup(ws4, 14)
title_block(ws4, "كشف التدفق النقدي الجنائي — 12 شهراً",
            "أين تختفي الأرباح؟ تتبع كل ريال شهرياً + فحص التسريبات الخفية — الفصلان 5 و8", 14)

r = 4
section(ws4, r, "التدفقات النقدية الشهرية — أدخل أرقامك في الخلايا ذات الإطار الذهبي", 14); r += 1
hdr = r
header_row(ws4, hdr, ["البند"] + MONTHS + ["الإجمالي"])
cash_in1 = r + 1
label_cell(ws4, cash_in1, 1, "المبيعات النقدية", bold=True)
for col in range(2, 14):
    input_cell(ws4, cash_in1, col, 270000 if col % 3 else 440000, '0')
res_cell(ws4, cash_in1, 14, f"=SUM(B{cash_in1}:M{cash_in1})", '0')
cash_in2 = cash_in1 + 1
label_cell(ws4, cash_in2, 1, "تحصيل المبيعات الآجلة", bold=True)
for col in range(2, 14):
    input_cell(ws4, cash_in2, col, 10000, '0')
res_cell(ws4, cash_in2, 14, f"=SUM(B{cash_in2}:M{cash_in2})", '0')
tot_in = cash_in2 + 1
label_cell(ws4, tot_in, 1, "إجمالي الداخل", bold=True)
res_cell(ws4, tot_in, 2, f"=B{cash_in1}+B{cash_in2}", '0')
for col in range(3, 14):
    cl = get_column_letter(col)
    res_cell(ws4, tot_in, col, f"={cl}{cash_in1}+{cl}{cash_in2}", '0')
res_cell(ws4, tot_in, 14, f"=SUM(B{tot_in}:M{tot_in})", '0')
for rr in (cash_in1, cash_in2, tot_in):
    ws4.row_dimensions[rr].height = 24

out_start = tot_in + 2
outflows = [
    ("شراء المواد الأولية", 105000),
    ("الرواتب والأجور", 88000),
    ("الإيجار", 40000),
    ("المرافق (كهرباء/ماء/غاز)", 12000),
    ("التسويق والدعاية", 8000),
    ("الصيانة", 6000),
    ("المصاريف الحكومية (إقامات/تأمين/تراخيص)", 15000),
    ("مصروفات أخرى", 9000),
]
for i, (name, val) in enumerate(outflows):
    rr = out_start + i
    label_cell(ws4, rr, 1, name, alt=(i % 2 == 1))
    for col in range(2, 14):
        input_cell(ws4, rr, col, val, '0')
    res_cell(ws4, rr, 14, f"=SUM(B{rr}:M{rr})", '0')
    ws4.row_dimensions[rr].height = 22
out_last = out_start + len(outflows) - 1
tot_out = out_last + 1
label_cell(ws4, tot_out, 1, "إجمالي الخارج", bold=True)
for col in range(2, 14):
    cl = get_column_letter(col)
    res_cell(ws4, tot_out, col, f"=SUM({cl}{out_start}:{cl}{out_last})", '0')
res_cell(ws4, tot_out, 14, f"=SUM(B{tot_out}:M{tot_out})", '0')
ws4.row_dimensions[tot_out].height = 26

net_r = tot_out + 2
label_cell(ws4, net_r, 1, "صافي التدفق", bold=True)
for col in range(2, 14):
    cl = get_column_letter(col)
    res_cell(ws4, net_r, col, f"={cl}{tot_in}-{cl}{tot_out}", '0')
res_cell(ws4, net_r, 14, f"=N{tot_in}-N{tot_out}", '0')
ws4.row_dimensions[net_r].height = 26

open_r = net_r + 1
label_cell(ws4, open_r, 1, "الرصيد الافتتاحي", bold=True)
input_cell(ws4, open_r, 2, 25000, '0')
for col in range(3, 14):
    cl = get_column_letter(col)
    prev = get_column_letter(col - 1)
    res_cell(ws4, open_r, col, f"={prev}{open_r+1}", '0')
res_cell(ws4, open_r, 14, f"=SUM(B{open_r}:M{open_r})", '0')
ws4.row_dimensions[open_r].height = 26

close_r = open_r + 1
label_cell(ws4, close_r, 1, "الرصيد الختامي", bold=True)
for col in range(2, 14):
    cl = get_column_letter(col)
    res_cell(ws4, close_r, col, f"={cl}{open_r}+{cl}{net_r}", '0')
res_cell(ws4, close_r, 14, f"=N{open_r}+N{net_r}", '0')
ws4.row_dimensions[close_r].height = 26

r = close_r + 2
section(ws4, r, "فحص التسريبات الخفية (Forensic Leak Scan)", 14); r += 1
leak_rows = [
    ("نسبة المواد الأولية", f"=IF(N{tot_in}>0,N{out_start}/N{tot_in},0)", 0.40, "المواد > 40% من الإيرادات"),
    ("نسبة العمالة", f"=IF(N{tot_in}>0,N{out_start+1}/N{tot_in},0)", 0.30, "العمالة > 30%"),
    ("نسبة التكلفة الأولية", f"=IF(N{tot_in}>0,(N{out_start}+N{out_start+1})/N{tot_in},0)", 0.65, "المواد + العمالة > 65%"),
    ("نسبة الإيجار", f"=IF(N{tot_in}>0,N{out_start+2}/N{tot_in},0)", 0.10, "الإيجار > 10%"),
    ("نسبة التسويق", f"=IF(N{tot_in}>0,N{out_start+4}/N{tot_in},0)", 0.05, "التسويق > 5%"),
]
leak_hdr = r
header_row(ws4, leak_hdr, ["المؤشر", "القيمة الفعلية", "الحد الأقصى", "الحالة", "الشرح"])
for i, (name, frm, limit, desc) in enumerate(leak_rows):
    rr = leak_hdr + 1 + i
    label_cell(ws4, rr, 1, name, bold=True, alt=(i % 2 == 1))
    res_cell(ws4, rr, 2, frm, '0.0%')
    res_cell(ws4, rr, 3, limit, '0.0%')
    res_cell(ws4, rr, 4, f'=IF(B{rr}>C{rr},"تسريب!","سليم")')
    ws4.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=7)
    c = ws4.cell(rr, 5, desc); c.font = F_LBL; c.alignment = AL_R
    ws4.row_dimensions[rr].height = 24
leak_last = leak_hdr + len(leak_rows)
r = leak_last + 2
label_cell(ws4, r, 1, "أدنى رصيد خلال العام (ر.س)", bold=True)
res_cell(ws4, r, 2, f"=MIN(B{close_r}:M{close_r})", '0')
ws4.row_dimensions[r].height = 26
r += 1
label_cell(ws4, r, 1, "عدد أشهر الرصيد السالب", bold=True)
res_cell(ws4, r, 2, f"=COUNTIF(B{close_r}:M{close_r},\"<0\")", '0')
ws4.row_dimensions[r].height = 26
r += 1

widths(ws4, [38] + [13] * 12 + [14])
autofit(ws4, min_w=11)
unlock4 = {(cash_in1, c) for c in range(2, 14)} | {(cash_in2, c) for c in range(2, 14)}
for rr in range(out_start, out_last + 1):
    unlock4 |= {(rr, c) for c in range(2, 14)}
unlock4 |= {(open_r, 2)}
lock_all(ws4, r, 14, unlock=unlock4)
ws4.conditional_formatting.add(f"B{close_r}:M{close_r}",
                               CellIsRule(operator="lessThan", formula=["0"],
                                          fill=PatternFill("solid", fgColor="FFC7CE"),
                                          font=Font(name="Calibri", color="9C0006", bold=True)))
ws4.conditional_formatting.add(f"D{leak_hdr+1}:D{leak_last}",
                               FormulaRule(formula=[f'$B{leak_hdr+1}>$C{leak_hdr+1}'],
                                           fill=PatternFill("solid", fgColor="FFC7CE"),
                                           font=Font(name="Calibri", color="9C0006", bold=True)))
ws4.page_setup.orientation = "landscape"
print("Sheet 4 done")

# ============================================================
# SHEET 5 — OPERATIONAL CHECKLISTS (Ch8)
# ============================================================
ws5 = wb.create_sheet("قوائم التشغيل")
base_setup(ws5, 4)
title_block(ws5, "قوائم التشغيل اليومية — Operational Checklists",
            "قائمة الفتح والتحضير وقائمة الإغلاق — جاهزة للتصدير PDF — الفصل الثامن", 4)

opening = [
    "تأكد من وصول جميع الموظفين وتسجيل الدوام",
    "فحص نظافة قاعة الطعام والمطابخ والمخازن",
    "معايرة موازين المطبخ وضبط درجات الثلاجات",
    "تأكد من توفر المواد الأساسية في المخزن",
    "تحضير وتجهيز قائمة اليوم والمقبلات",
    "اختبار أجهزة الدفع ونظام الـ POS",
    "فحص أنظمة سلامة الطعام وتاريخ الصلاحية",
    "تجهيز صالة الطعام والطاولات",
    "مراجعة الحجوزات والطلبات الخاصة",
    "اجتماع توجيهي سريع قبل الفتح",
]
closing = [
    "إغلاق نظام الـ POS وتسوية الفواتير اليومية",
    "مراجعة صندوق النقد وتسجيل إيداع اليوم",
    "حفظ المواد القابلة للتلف بدرجة حرارة صحيحة",
    "تنظيف المطبخ وطاولات التحضير وتعقيمها",
    "تفريغ مخلفات الطعام في أماكنها المخصصة",
    "إطفاء الغاز والأجهزة غير الضرورية",
    "تأمين الأبواب والنوافذ وتفعيل الإنذار",
    "تسجيل ملاحظات اليوم وقائمة المشتريات للغد",
    "مراجعة استهلاك المياه والكهرباء",
    "تأكيد جدول دوام الغد مع الفريق",
]


def checklist(ws, start_section, items, label):
    r = start_section
    section(ws, r, label, 4); r += 1
    header_row(ws, r, ["البند", "المسؤول", "مكتمل؟", "ملاحظات"])
    hdr = r
    first = r + 1
    for i, it in enumerate(items):
        rr = first + i
        label_cell(ws, rr, 1, it, alt=(i % 2 == 1))
        input_cell(ws, rr, 2, "الموظف", None)
        input_cell(ws, rr, 3, "لا")
        input_cell(ws, rr, 4, None)
        ws.row_dimensions[rr].height = 24
    last = first + len(items) - 1
    pct = last + 1
    label_cell(ws, pct, 1, "نسبة الإنجاز", bold=True)
    res_cell(ws, pct, 3, f"=COUNTIF(C{first}:C{last},\"نعم\")/COUNTA(C{first}:C{last})", '0%')
    ws.row_dimensions[pct].height = 26
    dv = DataValidation(type="list", formula1='"نعم,لا"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"C{first}:C{last}")
    ws.conditional_formatting.add(f"A{hdr}:D{pct}",
                                  FormulaRule(formula=[f'$C{first}="نعم"'],
                                              fill=PatternFill("solid", fgColor="C6EFCE"),
                                              font=Font(name="Calibri", color="006100")))
    return pct + 1


r = 4
r = checklist(ws5, r, opening, "قائمة الفتح الصباحية") + 1
r = checklist(ws5, r, closing, "قائمة الإغلاق المسائية")
widths(ws5, [60, 16, 14, 30])
autofit(ws5)
lock_all(ws5, r, 4,
         unlock={(rr, c) for rr in range(5, r + 1) for c in (2, 3, 4)})
ws5.page_setup.orientation = "portrait"
print("Sheet 5 done")

# ---------- save ----------
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("Saved:", OUT)
