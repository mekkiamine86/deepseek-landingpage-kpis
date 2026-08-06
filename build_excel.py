# -*- coding: utf-8 -*-
"""
Saudi Restaurant Intelligence 2026 — Excel Workbook Generator
Premium editorial dark-gold theme, Arabic financial terminology,
formula cells locked, only input cells open.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# ---------- Theme palette ----------
DEEP_GREEN   = "0A1F14"
DEEP_GREEN2  = "0E2B1C"
GOLD         = "D4AF37"
GOLD_LIGHT   = "F2E2A8"
IVORY        = "F7F1E0"
WHITE        = "FFFFFF"
DIM_TEXT     = "B8C4B8"
INPUT_BG     = "1A3A30"
INPUT_EDGE   = "D4AF37"
GREEN_GOOD   = "2ECC71"
RED          = "C0392B"
GRID         = "1E3A2C"

thin = Side(style="thin", color=GRID)
gold_border = Border(left=Side(style="thin", color=GOLD), right=Side(style="thin", color=GOLD),
                     top=Side(style="thin", color=GOLD), bottom=Side(style="thin", color=GOLD))

f_title    = Font(name="Calibri", size=18, bold=True, color=GOLD)
f_sub      = Font(name="Calibri", size=11, color=IVORY)
f_hdr      = Font(name="Calibri", size=11, bold=True, color=DEEP_GREEN)
f_section  = Font(name="Calibri", size=12, bold=True, color=GOLD_LIGHT)
f_lbl      = Font(name="Calibri", size=10, color=IVORY)
f_lbl_bold = Font(name="Calibri", size=10, bold=True, color=WHITE)
f_input    = Font(name="Calibri", size=10, color=WHITE)
f_result   = Font(name="Calibri", size=10, bold=True, color=GOLD_LIGHT)
f_small    = Font(name="Calibri", size=8.5, italic=True, color=DIM_TEXT)

fill_panel   = PatternFill("solid", fgColor=DEEP_GREEN2)
fill_hdr     = PatternFill("solid", fgColor=GOLD)
fill_input   = PatternFill("solid", fgColor=INPUT_BG)
fill_section = PatternFill("solid", fgColor=DEEP_GREEN)
fill_ok      = PatternFill("solid", fgColor=GREEN_GOOD)
fill_bad     = PatternFill("solid", fgColor=RED)

align_r = Alignment(horizontal="right", vertical="center")
align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_input(cell, wrap=False):
    cell.fill = fill_input
    cell.font = f_input
    cell.border = gold_border
    cell.protection = Protection(locked=False)
    cell.alignment = align_c if wrap else Alignment(horizontal="center", vertical="center")


def style_result(cell):
    cell.fill = fill_panel
    cell.font = f_result
    cell.border = Border(left=Side(style="thin", color=GOLD), right=Side(style="thin", color=GOLD),
                         top=Side(style="thin", color=GOLD), bottom=Side(style="thin", color=GOLD))
    cell.alignment = align_c


def style_label(cell, bold=False):
    cell.font = f_lbl_bold if bold else f_lbl
    cell.fill = fill_panel
    cell.alignment = align_r
    cell.border = Border(bottom=Side(style="thin", color=GRID))


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = f_title
    c.alignment = align_r
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = f_sub
    c2.alignment = align_r


def header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = f_hdr
        c.fill = fill_hdr
        c.alignment = align_c
        c.border = Border(bottom=Side(style="medium", color=DEEP_GREEN))
    ws.row_dimensions[row].height = 26


def section(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = f_section
    c.fill = fill_section
    c.alignment = align_r
    ws.row_dimensions[row].height = 22


wb = Workbook()
months = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
          "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]

# =========================================================
# SHEET 1 — README
# =========================================================
ws = wb.active
ws.title = "نظرة عامة"
ws.sheet_view.showGridLines = False
title_block(ws, "استخبارات المطاعم السعودية 2026",
            "باقة الأصول التشغيلية المالية — مصممة للبيئة التنظيمية السعودية (الزكاة، ضريبة القيمة المضافة 15%، التأمينات GOSI، منصة Qiwa)", 6)
ws.row_dimensions[3].height = 10

overview = [
    ("1. تتبع الزكاة وضريبة القيمة المضافة", "مؤشر ذكي للاحتياطي الضريبي 15% ومخصص الزكاة 2.5% وفق اشتراطات هيئة الزكاة والضريبة والجمارك (ZATCA)."),
    ("2. درع التكلفة الأولية والهدر", "رقابة تكلفة الطبق (Plate Costing) وهدر المواد الأولية بأحجام مناسبة للمطبخ السعودي والعربي."),
    ("3. مصفوفة تكلفة العمالة السعودية", "احتساب اشتراكات GOSI ورسوم المقيم و Muqeem/Qiwa وتأثيرها على الربحية الشهرية."),
    ("4. لوحة المؤشرات التنفيذية", "مخططات احترافية تفاعلية تعرض الإيرادات مقابل نقطة التعادل بضغطة واحدة."),
]
r = 5
for t, d in overview:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1, value=t); c.font = f_lbl_bold; c.alignment = align_r
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    c2 = ws.cell(row=r, column=3, value=d); c2.font = f_lbl; c2.alignment = align_r
    ws.row_dimensions[r].height = 26
    r += 1

r += 1
section(ws, r, "طريقة الاستخدام", 6); r += 1
for u in [
    "1. ابدأ من أوراق الإدخال: أدخل أرقامك في الخلايا ذات الإطار الذهبي فقط.",
    "2. كل خلايا الصيغ محمية ولا يمكن تعديلها (الحماية مفعّلة).",
    "3. افتح لوحة المؤشرات لقراءة النتيجة التنفيذية ومقارنة الإيراد بنقطة التعادل.",
    "4. تم بناء النموذج وفق معدلات 2026: ضريبة القيمة المضافة 15%، الزكاة 2.5%، اشتراكات GOSI.",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value=u); c.font = f_lbl; c.alignment = align_r
    ws.row_dimensions[r].height = 20
    r += 1

r += 1
section(ws, r, "إخلاء مسؤولية", 6); r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = ws.cell(row=r, column=1, value="المعدلات والاشتراطات واردة وفق الضوابط المنشورة لعام 2026. يُنصح بمراجعة مستشار زكاة وضريبة ومختص شؤون موظفين معتمد قبل التسوية النهائية. نموذج داخلي للإدارة المالية ولا يُغني عن الفوترة الإلكترونية ZATCA.")
c.font = f_small; c.alignment = align_r
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].width = 24

# =========================================================
# SHEET 2 — الزكاة وضريبة القيمة المضافة
# =========================================================
ws2 = wb.create_sheet("الزكاة وضريبة القيمة المضافة")
ws2.sheet_view.showGridLines = False
title_block(ws2, "مؤشر الزكاة وضريبة القيمة المضافة 15%",
            "احتياطي ضريبي تلقائي وفق اشتراطات ZATCA — أدخل المبيعات شاملة الضريبة", 8)
r = 4
section(ws2, r, "المدخلات العامة", 8); r += 1
style_label(ws2.cell(row=r, column=1, value="معدل ضريبة القيمة المضافة (السعودية)"))
style_input(ws2.cell(row=r, column=2, value=0.15)); r += 1
style_label(ws2.cell(row=r, column=1, value="معدل الزكاة (وعاء الزكاة المقدر)"))
style_input(ws2.cell(row=r, column=2, value=0.025))
VAT_RATE_ROW = 5
ZAKAT_RATE_ROW = 6
r += 2

section(ws2, r, "الجدول الشهري", 8); r += 1
headers2 = ["الشهر", "المبيعات شاملة الضريبة (ر.س)", "المبيعات الصافية (ر.س)",
            "ضريبة القيمة المضافة المحصلة", "ضريبة المدخلات (مشتريات)", "صافي ضريبة القيمة المضافة",
            "وعاء الزكاة (ر.س)", "مخصص الزكاة"]
header_row(ws2, r, headers2)
start2 = r + 1
sample_sales = [320000, 355000, 348000, 305000, 520000, 610000,
                480000, 470000, 505000, 560000, 445000, 620000]
sample_purchases = [120000, 130000, 128000, 110000, 190000, 215000,
                    170000, 168000, 180000, 198000, 158000, 220000]
sample_zakat_base = [180000, 200000, 195000, 175000, 260000, 290000,
                     235000, 230000, 250000, 275000, 220000, 295000]
for i, m in enumerate(months):
    rr = start2 + i
    ws2.cell(row=rr, column=1, value=m).font = f_lbl_bold
    ws2.cell(row=rr, column=1).fill = fill_panel
    ws2.cell(row=rr, column=1).alignment = align_r
    style_input(ws2.cell(row=rr, column=2, value=sample_sales[i]))
    ws2.cell(row=rr, column=3, value=f"=B{rr}/(1+$B${VAT_RATE_ROW})")
    ws2.cell(row=rr, column=4, value=f"=B{rr}-C{rr}")
    style_input(ws2.cell(row=rr, column=5, value=sample_purchases[i]))
    ws2.cell(row=rr, column=6, value=f"=MAX(0,D{rr}-E{rr})")
    style_input(ws2.cell(row=rr, column=7, value=sample_zakat_base[i]))
    ws2.cell(row=rr, column=8, value=f"=G{rr}*$B${ZAKAT_RATE_ROW}")
    for col in (3, 4, 6, 8):
        style_result(ws2.cell(row=rr, column=col))
    ws2.row_dimensions[rr].height = 18
last2 = start2 + 11
tr = last2 + 1
ws2.cell(row=tr, column=1, value="الإجمالي السنوي").font = f_lbl_bold
ws2.cell(row=tr, column=1).fill = fill_hdr
for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H")):
    cell = ws2.cell(row=tr, column=col, value=f"=SUM({letter}{start2}:{letter}{last2})")
    style_result(cell)
    cell.font = f_hdr
    cell.fill = fill_hdr
for col, w in zip(range(1, 9), [12, 24, 20, 22, 22, 20, 18, 16]):
    ws2.column_dimensions[get_column_letter(col)].width = w

# =========================================================
# SHEET 3 — درع التكلفة الأولية والهدر
# =========================================================
ws3 = wb.create_sheet("درع التكلفة والهدر")
ws3.sheet_view.showGridLines = False
title_block(ws3, "درع التكلفة الأولية والهدر (Prime Cost Shield)",
            "تسعير الطبق وفق أحجام المطبخ السعودي والعربي + رقابة هدر المواد الأولية", 7)
r = 4
section(ws3, r, "تكلفة الطبق — كشف تسعير المنيو", 7); r += 1
headers3 = ["الطبق", "تكلفة المكونات (ر.س)", "سعر البيع شامل الضريبة (ر.س)",
            "سعر البيع الصافي (ر.س)", "هامش الربح الخام (ر.س)", "نسبة تكلفة المكونات", "الحالة"]
header_row(ws3, r, headers3)
start3 = r + 1
dishes = [
    ("كبسة لحم", 14.50, 55.0),
    ("مندي دجاج", 10.80, 42.0),
    ("جريش", 9.20, 36.0),
    ("مقبلات مشكلة", 6.50, 28.0),
    ("كوزي", 22.00, 78.0),
    ("حلى كنافة", 5.00, 24.0),
    ("بروستد (قطعة)", 4.20, 18.0),
]
for i, (dish, cost, price) in enumerate(dishes):
    rr = start3 + i
    ws3.cell(row=rr, column=1, value=dish).font = f_lbl_bold
    ws3.cell(row=rr, column=1).fill = fill_panel
    ws3.cell(row=rr, column=1).alignment = align_r
    style_input(ws3.cell(row=rr, column=2, value=cost))
    style_input(ws3.cell(row=rr, column=3, value=price))
    ws3.cell(row=rr, column=4, value=f"=C{rr}/(1+$B${VAT_RATE_ROW})")
    ws3.cell(row=rr, column=5, value=f"=D{rr}-B{rr}")
    ws3.cell(row=rr, column=6, value=f"=B{rr}/D{rr}")
    ws3.cell(row=rr, column=7, value=f'=IF(F{rr}<=0.30,"ممتاز",IF(F{rr}<=0.38,"مقبول","إعادة تسعير"))')
    for col in (4, 5, 6, 7):
        style_result(ws3.cell(row=rr, column=col))
    ws3.row_dimensions[rr].height = 18
last3 = start3 + len(dishes) - 1
r = last3 + 2

section(ws3, r, "مؤشر هدر المواد الأولية الشهري", 7); r += 1
headers3b = ["الشهر", "تكلفة المبيعات (COGS)", "المبيعات الصافية", "قيمة الهدر المسجلة (ر.س)",
             "نسبة الهدر", "الحد المسموح", "الحالة"]
header_row(ws3, r, headers3b)
start3b = r + 1
sample_cogs = [108000, 118000, 114000, 98000, 168000, 190000,
               156000, 152000, 163000, 180000, 142000, 195000]
sample_waste = [4000, 5200, 4800, 6100, 7900, 8200, 6600, 5900, 7200, 8400, 6300, 9800]
for i, m in enumerate(months):
    rr = start3b + i
    ws3.cell(row=rr, column=1, value=m).font = f_lbl_bold
    ws3.cell(row=rr, column=1).fill = fill_panel
    ws3.cell(row=rr, column=1).alignment = align_r
    style_input(ws3.cell(row=rr, column=2, value=sample_cogs[i]))
    ws3.cell(row=rr, column=3, value=f"=B{rr}/(1+$B${VAT_RATE_ROW})")
    style_input(ws3.cell(row=rr, column=4, value=sample_waste[i]))
    ws3.cell(row=rr, column=5, value=f"=E{rr}/C{rr}")
    ws3.cell(row=rr, column=6, value=0.03)
    ws3.cell(row=rr, column=7, value=f'=IF(E{rr}/C{rr}<=0.03,"ضمن الحد","خارج الحد")')
    for col in (3, 5, 7):
        style_result(ws3.cell(row=rr, column=col))
    style_input(ws3.cell(row=rr, column=6, value=0.03), wrap=False)
    ws3.row_dimensions[rr].height = 18
for col, w in zip(range(1, 8), [20, 22, 20, 22, 16, 16, 18]):
    ws3.column_dimensions[get_column_letter(col)].width = w

# =========================================================
# SHEET 4 — مصفوفة تكلفة العمالة السعودية
# =========================================================
ws4 = wb.create_sheet("مصفوفة تكلفة العمالة")
ws4.sheet_view.showGridLines = False
title_block(ws4, "مصفوفة تكلفة العمالة السعودية 2026",
            "اشتراكات GOSI + رسوم المقيم (Muqeem) وتكلفة Qiwa — أثر التكلفة الكلية على الربحية", 9)
r = 4
section(ws4, r, "معدلات واشتراكات عام 2026", 9); r += 1
style_label(ws4.cell(row=r, column=1, value="اشتراك GOSI — حصة الموظف (سعودي)"))
style_input(ws4.cell(row=r, column=2, value=0.0975)); GOSI_EMP_ROW = r; r += 1
style_label(ws4.cell(row=r, column=1, value="اشتراك GOSI — حصة صاحب العمل (سعودي)"))
style_input(ws4.cell(row=r, column=2, value=0.1175)); GOSI_EMP_ER_ROW = r; r += 1
style_label(ws4.cell(row=r, column=1, value="التأمين ضد مخاطر العمل (سعودي + وافد)"))
style_input(ws4.cell(row=r, column=2, value=0.02)); GOSI_OC_ROW = r; r += 1
style_label(ws4.cell(row=r, column=1, value="اشتراك GOSI — صاحب العمل (وافد) مخاطر العمل فقط"))
style_input(ws4.cell(row=r, column=2, value=0.02)); r += 1
style_label(ws4.cell(row=r, column=1, value="رسوم رخصة العمل الشهرية (Qiwa) لكل وافد (ر.س)"))
style_input(ws4.cell(row=r, column=2, value=125)); QIWA_FEE_ROW = r; r += 1
style_label(ws4.cell(row=r, column=1, value="رسوم الإقامة (Muqeem) الشهرية لكل وافد (ر.س)"))
style_input(ws4.cell(row=r, column=2, value=200)); MUQEEM_FEE_ROW = r
r += 2

section(ws4, r, "سجل الموظفين — التكلفة الفعلية على صاحب العمل", 9); r += 1
headers4 = ["الاسم", "الجنسية", "الأساسي (ر.س)", "البدلات (ر.س)", "إجمالي الأجر (ر.س)",
            "GOSI صاحب العمل", "رسوم Qiwa + Muqeem", "التكلفة الشهرية الكلية", "نسبة التكلفة"]
header_row(ws4, r, headers4)
start4 = r + 1
staff = [
    ("شيف عربي", "سعودي", 9000, 1500),
    ("شيف مساعد", "وافد", 4500, 700),
    ("نادل أول", "وافد", 3200, 500),
    ("كاشير", "سعودي", 5000, 800),
    ("مشرف قاعة", "سعودي", 6500, 1000),
    ("عامل مطبخ", "وافد", 2800, 400),
]
for i, (name, nat, base, allow) in enumerate(staff):
    rr = start4 + i
    ws4.cell(row=rr, column=1, value=name).font = f_lbl_bold
    ws4.cell(row=rr, column=1).fill = fill_panel
    ws4.cell(row=rr, column=1).alignment = align_r
    ws4.cell(row=rr, column=2, value=nat).font = f_lbl
    ws4.cell(row=rr, column=2).fill = fill_panel
    ws4.cell(row=rr, column=2).alignment = align_c
    style_input(ws4.cell(row=rr, column=3, value=base))
    style_input(ws4.cell(row=rr, column=4, value=allow))
    ws4.cell(row=rr, column=5, value=f"=C{rr}+D{rr}")
    ws4.cell(row=rr, column=6, value=f'=IF(B{rr}="سعودي",E{rr}*($B${GOSI_EMP_ER_ROW}+$B${GOSI_OC_ROW}),E{rr}*$B${GOSI_OC_ROW})')
    ws4.cell(row=rr, column=7, value=f'=IF(B{rr}="سعودي",0,($B${QIWA_FEE_ROW}+$B${MUQEEM_FEE_ROW}))')
    ws4.cell(row=rr, column=8, value=f"=E{rr}+F{rr}+G{rr}")
    ws4.cell(row=rr, column=9, value=f'=IF(SUM($E${start4}:$E${start4+len(staff)-1})>0,H{rr}/SUM($E${start4}:$E${start4+len(staff)-1}),0)')
    for col in (5, 6, 7, 8, 9):
        style_result(ws4.cell(row=rr, column=col))
    ws4.row_dimensions[rr].height = 18
last4 = start4 + len(staff) - 1
tr = last4 + 1
ws4.cell(row=tr, column=1, value="الإجمالي").font = f_lbl_bold
ws4.cell(row=tr, column=1).fill = fill_hdr
for col, letter in ((5, "E"), (6, "F"), (7, "G"), (8, "H")):
    cell = ws4.cell(row=tr, column=col, value=f"=SUM({letter}{start4}:{letter}{last4})")
    style_result(cell); cell.font = f_hdr; cell.fill = fill_hdr
ws4.cell(row=tr, column=9, value=f"=IF(H{tr}>0,1,0)")
ws4.cell(row=tr, column=9).font = f_hdr
ws4.cell(row=tr, column=9).fill = fill_hdr

for col, w in zip(range(1, 10), [18, 12, 14, 14, 18, 20, 22, 22, 16]):
    ws4.column_dimensions[get_column_letter(col)].width = w

# =========================================================
# SHEET 5 — لوحة المؤشرات التنفيذية
# =========================================================
ws5 = wb.create_sheet("لوحة المؤشرات")
ws5.sheet_view.showGridLines = False
title_block(ws5, "لوحة المؤشرات التنفيذية — الإيرادات مقابل نقطة التعادل",
            "اختر الشهر من القائمة، واقرأ الوضع المالي فوراً", 8)

# KPI block
r = 4
kpis = [
    ("الإيرادات الصافية للشهر (ر.س)", 6),
    ("التكلفة الأولية (Prime Cost)", 7),
    ("صافي الربح التشغيلي (ر.س)", 8),
    ("نقطة التعادل الشهرية (ر.س)", 9),
    ("هامش الأمان (ر.س)", 10),
    ("نسبة التكلفة الأولية", 11),
]
# Build a small data table for the selected month (rows 6-11 col A label, col B value)
sel_row = 5  # month selector lives here
style_label(ws5.cell(row=sel_row, column=1, value="الشهر المحدد"))
dv = DataValidation(type="list", formula1='"يناير,فبراير,مارس,أبريل,مايو,يونيو,يوليو,أغسطس,سبتمبر,أكتوبر,نوفمبر,ديسمبر"', allow_blank=True)
ws5.add_data_validation(dv)
csel = ws5.cell(row=sel_row, column=2, value="يناير")
dv.add(csel)
style_input(csel, wrap=False)

# Data for selected month (input columns referencing sheet2 sales)
r = 7
ws5.cell(row=r, column=1, value="الإيرادات شاملة الضريبة (ر.س)").font = f_lbl_bold
ws5.cell(row=r, column=1).fill = fill_panel
ws5.cell(row=r, column=1).alignment = align_r
ws5.cell(row=r, column=2, value=f'=INDEX(\'الزكاة وضريبة القيمة المضافة\'!$B${start2}:$B${last2},MATCH($B$5,\'الزكاة وضريبة القيمة المضافة\'!$A${start2}:$A${last2},0))')
style_result(ws5.cell(row=r, column=2)); rev_row = r
r += 1
ws5.cell(row=r, column=1, value="التكلفة الأولية (مواد + عمالة)").font = f_lbl_bold
ws5.cell(row=r, column=1).fill = fill_panel
ws5.cell(row=r, column=1).alignment = align_r
ws5.cell(row=r, column=2, value=f'=INDEX(\'درع التكلفة والهدر\'!$B${start3b}:$B${last3b},MATCH($B$5,\'درع التكلفة والهدر\'!$A${start3b}:$A${last3b},0))+$H${tr}')
style_result(ws5.cell(row=r, column=2)); prime_row = r
r += 1
ws5.cell(row=r, column=1, value="المصاريف التشغيلية الثابتة (ر.س)").font = f_lbl_bold
ws5.cell(row=r, column=1).fill = fill_panel
ws5.cell(row=r, column=1).alignment = align_r
c = ws5.cell(row=r, column=2, value=65000); style_input(c, wrap=False); fixed_row = r
r += 1
ws5.cell(row=r, column=1, value="نقطة التعادل الشهرية (ر.س)").font = f_lbl_bold
ws5.cell(row=r, column=1).fill = fill_panel
ws5.cell(row=r, column=1).alignment = align_r
ws5.cell(row=r, column=2, value=f"=IF({rev_row}>0,{fixed_row}/(1-{prime_row}/{rev_row}),0)")
style_result(ws5.cell(row=r, column=2)); bep_row = r
r += 1
ws5.cell(row=r, column=1, value="صافي الربح التشغيلي (ر.س)").font = f_lbl_bold
ws5.cell(row=r, column=1).fill = fill_panel
ws5.cell(row=r, column=1).alignment = align_r
ws5.cell(row=r, column=2, value=f"={rev_row}-{prime_row}-{fixed_row}")
style_result(ws5.cell(row=r, column=2)); profit_row = r
r += 1
ws5.cell(row=r, column=1, value="هامش الأمان (ر.س)").font = f_lbl_bold
ws5.cell(row=r, column=1).fill = fill_panel
ws5.cell(row=r, column=1).alignment = align_r
ws5.cell(row=r, column=2, value=f"={rev_row}-{bep_row}")
style_result(ws5.cell(row=r, column=2)); safety_row = r
r += 2

section(ws5, r, "التوزيع الشهري — الإيرادات مقابل نقطة التعادل", 8); r += 1
# mini table for chart: month, revenue, breakeven
ws5.cell(row=r, column=1, value="الشهر").font = f_lbl_bold
ws5.cell(row=r, column=2, value="الإيرادات").font = f_lbl_bold
ws5.cell(row=r, column=3, value="نقطة التعادل").font = f_lbl_bold
for ccell in (ws5.cell(row=r, column=1), ws5.cell(row=r, column=2), ws5.cell(row=r, column=3)):
    ccell.fill = fill_hdr
chart_start = r + 1
for i, m in enumerate(months):
    rr = chart_start + i
    ws5.cell(row=rr, column=1, value=m).font = f_lbl
    ws5.cell(row=rr, column=1).fill = fill_panel
    ws5.cell(row=rr, column=2, value=f'=INDEX(\'الزكاة وضريبة القيمة المضافة\'!$C${start2}:$C${last2},{i+1})')
    ws5.cell(row=rr, column=3, value=f"={fixed_row}/(1-{prime_row}/{rev_row})")
    for col in (2, 3):
        style_result(ws5.cell(row=rr, column=col))
chart_last = chart_start + 11

# Charts
ch1 = LineChart()
ch1.title = "الإيرادات الصافية مقابل نقطة التعادل"
ch1.style = 13
ch1.height = 9
ch1.width = 20
data = Reference(ws5, min_col=2, min_row=r, max_col=3, max_row=chart_last)
cats = Reference(ws5, min_col=1, min_row=chart_start, max_row=chart_last)
ch1.add_data(data, titles_from_data=True)
ch1.set_categories(cats)
ch1.y_axis.title = "ر.س"
ch1.x_axis.title = "الشهر"
ws5.add_chart(ch1, "E7")

ch2 = BarChart()
ch2.type = "col"
ch2.title = "مقارنة ربحية كل شهر"
ch2.style = 13
ch2.height = 9
ch2.width = 20
data2 = Reference(ws5, min_col=2, min_row=r, max_col=2, max_row=chart_last)
ch2.add_data(data2, titles_from_data=True)
ch2.set_categories(cats)
ws5.add_chart(ch2, "E20")

# Chart series colors
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.fill import PatternFillProperties
for s in ch1.series:
    s.graphicalProperties.line.solidFill = GOLD if s.idx == 0 else RED

for col, w in zip(range(1, 9), [28, 20, 18, 18, 18, 18, 18, 18]):
    ws5.column_dimensions[get_column_letter(col)].width = w

# =========================================================
# Protection pass — lock everything except unlocked inputs
# =========================================================
for sheet in (ws2, ws3, ws4, ws5):
    sheet.protection.sheet = True
    sheet.protection.enable()

out = r"C:\Users\DATA NET\Desktop\deepseek-landingpage-kpis\public\Saudi-Restaurant-Intelligence-2026.xlsx"
wb.save(out)
print("SAVED:", out)
